"""
Extrae datos del Excel 'data/reporte.xlsx' y genera data.js
con todas las estadísticas preprocesadas para el dashboard.

Script Robusto: Detecta dinámicamente el número de controles y la columna Total.
"""
import openpyxl
import json
import math
import os
from datetime import datetime

# Rutas relativas
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, 'data', 'reporte.xlsx')
OUTPUT_PATH = os.path.join(SCRIPT_DIR, 'data.js')

def extract_section(ws, section_name):
    """Extract student data from a worksheet dynamically."""
    # 1. Identificar estructura de columnas en la fila 2 (headers)
    headers = [cell.value for cell in ws[2]]
    
    try:
        total_idx = headers.index("Total")
    except ValueError:
        # Si no hay columna "Total", buscamos la última con datos en la fila 2
        total_idx = len(headers) - 1
        for i, h in enumerate(headers):
            if h is None:
                total_idx = i - 1
                break

    # Los controles empiezan en la columna F (índice 5)
    # y terminan justo antes del Total
    start_control_idx = 5
    num_controls_possible = total_idx - start_control_idx
    
    students = []
    max_active_control = 0
    
    # 2. Extraer datos de estudiantes (desde fila 3)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=total_idx + 1):
        vals = [cell.value for cell in row]
        if not vals or vals[0] is None:
            continue
        
        scores = []
        for i in range(num_controls_possible):
            v = vals[start_control_idx + i]
            score = float(v) if v is not None and str(v).replace('.','').isdigit() else 0
            scores.append(score)
            if score > 0:
                max_active_control = max(max_active_control, i + 1)
        
        students.append({
            'name': f"{vals[1]} {vals[2]}, {vals[3]}".strip() if len(vals) > 3 else "Sin Nombre",
            'id': str(vals[4]) if len(vals) > 4 and vals[4] else '',
            'scores': scores,
            'total': float(vals[total_idx]) if vals[total_idx] is not None else sum(scores)
        })

    # Ajustar a la cantidad de controles que realmente tienen alguna nota
    control_labels = [f"C{i+1}" for i in range(max_active_control)]
    for s in students:
        s['scores'] = s['scores'][:max_active_control]

    return {
        'name': section_name,
        'controlLabels': control_labels,
        'students': students,
        'numControls': max_active_control
    }

def compute_stats(section):
    """Compute all statistics for a section."""
    num_controls = section['numControls']
    students = section['students']
    n_students = len(students)
    max_per_control = 5

    stats = {
        'controlLabels': section['controlLabels'],
        'numStudents': n_students,
        'maxPerControl': max_per_control,
    }

    mins, maxs, avgs, avgs_responded, responded_counts = [], [], [], [], []
    cumulative_mins, cumulative_maxs, cumulative_avgs, cumulative_nota7 = [], [], [], []
    approval_pcts, boxplot_data, distributions = [], [], []

    cum_min, cum_max, cum_sum = 0, 0, 0

    for c in range(num_controls):
        scores = [s['scores'][c] for s in students]
        nonzero_scores = [s for s in scores if s > 0]

        c_min = min(scores) if scores else 0
        c_max = max(scores) if scores else 0
        c_avg = sum(scores) / n_students if n_students > 0 else 0
        c_avg_resp = sum(nonzero_scores) / len(nonzero_scores) if nonzero_scores else 0
        c_responded = len(nonzero_scores)
        c_approval = sum(1 for s in scores if s >= 3) / n_students * 100 if n_students > 0 else 0

        mins.append(round(c_min, 2))
        maxs.append(round(c_max, 2))
        avgs.append(round(c_avg, 2))
        avgs_responded.append(round(c_avg_resp, 2))
        responded_counts.append(c_responded)
        approval_pcts.append(round(c_approval, 1))

        cum_min += c_min
        cum_max += max_per_control
        cum_sum += c_avg
        cumulative_mins.append(round(cum_min, 2))
        cumulative_maxs.append(round(cum_max, 2))
        cumulative_avgs.append(round(cum_sum, 2))
        cumulative_nota7.append(round(cum_max * 0.8, 2))

        # Boxplot data
        sorted_scores = sorted(scores)
        def percentile(arr, p):
            if not arr: return 0
            k = (len(arr) - 1) * p
            f, c_val = math.floor(k), math.ceil(k)
            return arr[f] if f == c_val else arr[f] * (c_val-k) + arr[c_val] * (k-f)

        boxplot_data.append({
            'min': sorted_scores[0] if sorted_scores else 0,
            'q1': round(percentile(sorted_scores, 0.25), 2),
            'median': round(percentile(sorted_scores, 0.5), 2),
            'q3': round(percentile(sorted_scores, 0.75), 2),
            'max': sorted_scores[-1] if sorted_scores else 0,
            'mean': round(c_avg, 2)
        })

        dist = [0] * 6
        for s in scores:
            if 0 <= s <= 5: dist[int(s)] += 1
        distributions.append(dist)

    stats.update({
        'mins': mins, 'maxs': maxs, 'avgs': avgs, 'avgsResponded': avgs_responded,
        'respondedCounts': responded_counts, 'approvalPcts': approval_pcts,
        'cumulativeMins': cumulative_mins, 'cumulativeMaxs': cumulative_maxs,
        'cumulativeAvgs': cumulative_avgs, 'cumulativeNota7': cumulative_nota7,
        'boxplot': boxplot_data, 'distributions': distributions,
        'heatmap': {
            'names': [s['id'] if s['id'] else "Sin ID" for s in students],
            'scores': [s['scores'] for s in students],
            'totals': [s['total'] for s in students]
        }
    })
    
    ranked = sorted(students, key=lambda s: s['total'], reverse=True)
    stats['ranking'] = {
        'names': [s['id'] if s['id'] else "Sin ID" for s in ranked[:15]],
        'totals': [s['total'] for s in ranked[:15]]
    }
    return stats

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: No se encontró el Excel en: {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    stats1 = compute_stats(extract_section(wb[wb.sheetnames[0]], 'Sección 1'))
    stats2 = compute_stats(extract_section(wb[wb.sheetnames[1]], 'Sección 2'))

    combined = {
        'name': 'Ambas Secciones',
        'controlLabels': stats1['controlLabels'],
        'students': [], # No necesitamos la lista cruda aquí
        'numStudents': stats1['numStudents'] + stats2['numStudents'],
        'numControls': stats1['numControls'],
        'maxPerControl': 5
    }
    
    # Re-computar stats combinadas de forma simple (o podrías concatenar alumnos)
    # Por simplicidad, concatenamos alumnos para usar las mismas funciones
    all_students = []
    # Necesitamos reconstruir los objetos de estudiantes para el cálculo combinado
    # pero el extract_section ya lo hace. Vamos a simplificar el flujo:
    s1_full = extract_section(wb[wb.sheetnames[0]], 'Sección 1')
    s2_full = extract_section(wb[wb.sheetnames[1]], 'Sección 2')
    combined_full = {
        'name': 'Ambas Secciones',
        'controlLabels': s1_full['controlLabels'],
        'students': s1_full['students'] + s2_full['students'],
        'numControls': s1_full['numControls']
    }
    stats_combined = compute_stats(combined_full)

    today = datetime.now().strftime('%Y-%m-%d')
    data = {
        'section1': stats1, 'section2': stats2, 'combined': stats_combined,
        'sectionNames': ['Sección 1', 'Sección 2', 'Ambas Secciones'],
        'maxPerControl': 5, 'generatedAt': today
    }

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(f"const REPORT_DATA = {json.dumps(data, ensure_ascii=False, indent=2)};\n")

    print(f"Extraction complete: {stats_combined['numControls']} controls detected.")

if __name__ == '__main__':
    main()
